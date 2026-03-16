"""
Excel-Writer mit MyD Formatierung - v4.
Multi-Protokoll Support + INDEX/MATCH Formeln.
Unterstuetzt EN + DE Protokolle via _match_key.
Enthaelt Anleitung-Sheet + intelligente Auswertung.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

from config.theme import (
    EXCEL_ORANGE, EXCEL_ORANGE_LIGHT, EXCEL_EVEN_ROW, EXCEL_TOTAL_BG,
    EXCEL_RED, EXCEL_GREEN, EXCEL_YELLOW, EXCEL_WHITE, EXCEL_DARK, EXCEL_BG,
    EXCEL_SUCCESS_BG, EXCEL_ERROR_BG, EXCEL_WARN_BG
)


class MyDStyles:
    def __init__(self):
        self.title = Font(name='Arial', bold=True, color=EXCEL_ORANGE, size=16)
        self.subtitle = Font(name='Arial', bold=True, color=EXCEL_DARK, size=12)
        self.header = Font(name='Arial', bold=True, color=EXCEL_WHITE, size=10)
        self.data = Font(name='Arial', size=10)
        self.data_bold = Font(name='Arial', bold=True, size=10)
        self.kpi_value = Font(name='Arial', bold=True, color=EXCEL_ORANGE, size=22)
        self.kpi_label = Font(name='Arial', color=EXCEL_DARK, size=9)
        self.success = Font(name='Arial', bold=True, color=EXCEL_GREEN, size=11)
        self.error = Font(name='Arial', bold=True, color=EXCEL_RED, size=11)
        self.warning = Font(name='Arial', bold=True, color=EXCEL_YELLOW, size=11)
        self.total = Font(name='Arial', bold=True, color=EXCEL_ORANGE, size=11)
        self.formula_hint = Font(name='Arial', italic=True, color='999999', size=9)
        self.question = Font(name='Arial', italic=True, color=EXCEL_ORANGE, size=10)
        self.info_text = Font(name='Arial', color='555555', size=10)

        self.orange_fill = PatternFill(start_color=EXCEL_ORANGE, end_color=EXCEL_ORANGE, patternType='solid')
        self.dark_fill = PatternFill(start_color=EXCEL_DARK, end_color=EXCEL_DARK, patternType='solid')
        self.even_fill = PatternFill(start_color=EXCEL_EVEN_ROW, end_color=EXCEL_EVEN_ROW, patternType='solid')
        self.total_fill = PatternFill(start_color=EXCEL_TOTAL_BG, end_color=EXCEL_TOTAL_BG, patternType='solid')
        self.bg_fill = PatternFill(start_color=EXCEL_BG, end_color=EXCEL_BG, patternType='solid')
        self.red_fill = PatternFill(start_color=EXCEL_RED, end_color=EXCEL_RED, patternType='solid')
        self.success_bg = PatternFill(start_color=EXCEL_SUCCESS_BG, end_color=EXCEL_SUCCESS_BG, patternType='solid')
        self.error_bg = PatternFill(start_color=EXCEL_ERROR_BG, end_color=EXCEL_ERROR_BG, patternType='solid')
        self.warn_bg = PatternFill(start_color=EXCEL_WARN_BG, end_color=EXCEL_WARN_BG, patternType='solid')
        self.kpi_fill = PatternFill(start_color=EXCEL_WHITE, end_color=EXCEL_WHITE, patternType='solid')
        self.question_bg = PatternFill(start_color='FFF3E0', end_color='FFF3E0', patternType='solid')

        self.center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left = Alignment(horizontal='left', vertical='center')
        self.left_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True)
        self.right = Alignment(horizontal='right', vertical='center')

        self.thin = Border(
            left=Side(style='thin', color=EXCEL_ORANGE),
            right=Side(style='thin', color=EXCEL_ORANGE),
            top=Side(style='thin', color=EXCEL_ORANGE),
            bottom=Side(style='thin', color=EXCEL_ORANGE))
        self.kpi_border = Border(
            left=Side(style='thin', color=EXCEL_EVEN_ROW),
            right=Side(style='thin', color=EXCEL_EVEN_ROW),
            top=Side(style='thin', color=EXCEL_EVEN_ROW),
            bottom=Side(style='medium', color=EXCEL_ORANGE))


def create_excel(all_data, skipped_files, output_path, all_protocols=None):
    wb = Workbook()
    wb.remove(wb.active)
    s = MyDStyles()
    if all_protocols is None:
        all_protocols = []

    # Merged lookup ueber _match_key (nur erstes Key-Feld = Product/Material Number)
    merged_lookup = {}
    for proto in all_protocols:
        for row in proto['data']:
            mk = row.get('_match_key', '').strip()
            if mk:
                merged_lookup[mk] = row

    # --- Per-Key-Typ Statistiken berechnen ---
    per_key_stats = {}
    sn = 0
    for key_sig, info in all_data.items():
        sn += 1
        first_field = info['fields'][0]  # Product Number / Material Number

        # Unique Match-Keys fuer diesen Key-Typ (nur erstes Feld)
        key_matches = {}
        for row in info['rows']:
            pk = row.get(first_field, '').strip()
            if pk:
                if pk not in key_matches:
                    key_matches[pk] = []
                key_matches[pk].append(row)

        unique_keys = set(key_matches.keys())
        missing = unique_keys - set(merged_lookup.keys())
        matched = unique_keys - missing

        missing_by_file = {}
        for mk in missing:
            for row in key_matches[mk]:
                fname = row.get('_filename', '')
                if fname not in missing_by_file:
                    missing_by_file[fname] = []
                if mk not in missing_by_file[fname]:
                    missing_by_file[fname].append(mk)

        per_key_stats[key_sig] = {
            'key_num': sn,
            'first_field': first_field,
            'total_rows': len(info['rows']),
            'unique_keys': len(unique_keys),
            'matched': len(matched),
            'missing': len(missing),
            'has_duplicates': len(info['rows']) > len(unique_keys),
            'duplicate_count': len(info['rows']) - len(unique_keys),
            'missing_by_file': missing_by_file,
        }

    # Proto-Keys die in KEINEM Key-Typ vorkommen
    all_xml_first_keys = set()
    for key_sig, info in all_data.items():
        first_field = info['fields'][0]
        for row in info['rows']:
            pk = row.get(first_field, '').strip()
            if pk:
                all_xml_first_keys.add(pk)

    proto_only_keys = []
    for proto in all_protocols:
        for row in proto['data']:
            mk = row.get('_match_key', '').strip()
            if mk and mk not in all_xml_first_keys:
                proto_only_keys.append(mk)

    # Legacy missing_by_file fuer Dashboard (Gesamt ueber alle Key-Typen)
    missing_by_file_global = {}
    for ks_info in per_key_stats.values():
        for fname, keys in ks_info['missing_by_file'].items():
            if fname not in missing_by_file_global:
                missing_by_file_global[fname] = []
            missing_by_file_global[fname].extend(keys)

    _create_dashboard(wb, all_data, skipped_files, s, all_protocols,
                      missing_by_file_global, merged_lookup, proto_only_keys,
                      per_key_stats=per_key_stats)
    _create_anleitung(wb, s, all_data)

    proto_sheet_names = []
    for i, proto in enumerate(all_protocols):
        sname = "Protokoll" if len(all_protocols) == 1 else f"Proto_{i+1}"
        proto_sheet_names.append(sname)
        _create_protocol_sheet(wb, proto, s, sname)

    _create_data_sheets(wb, all_data, s)

    # --- EIN Vergleich-Sheet pro Key-Typ (pro Protokoll) ---
    for i, proto in enumerate(all_protocols):
        sn = 0
        for key_sig, info in all_data.items():
            sn += 1
            if len(all_protocols) == 1:
                vname = f"Vergl_{sn}"
            else:
                vname = f"Vergl_{i+1}_{sn}"
            _create_comparison_sheet(wb, info, sn, s, proto_sheet_names[i], proto, vname,
                                     per_key_stats.get(key_sig, {}))

    if skipped_files:
        _create_error_sheet(wb, skipped_files, s)

    wb.calculation = None
    wb.save(output_path)


def _create_dashboard(wb, all_data, skipped_files, s, all_protocols,
                      missing_by_file, merged_lookup, proto_only_keys,
                      per_key_stats=None):
    ws = wb.create_sheet(title="Uebersicht", index=0)
    for row in range(1, 120):
        for col in range(1, 12):
            ws.cell(row=row, column=col).fill = s.bg_fill
    for col in range(1, 10):
        ws.cell(row=1, column=col).fill = s.orange_fill
    ws.cell(row=2, column=2, value="SAP Migration Checker").font = s.title
    ws.merge_cells('B2:H2')
    ws.cell(row=3, column=2, value="Automatisierte Pruefung der Migrationsergebnisse").font = Font(name='Arial', italic=True, color=EXCEL_DARK, size=10)
    ws.merge_cells('B3:H3')
    for col in range(1, 10):
        ws.cell(row=4, column=col).border = Border(bottom=Side(style='medium', color=EXCEL_ORANGE))

    # --- Grundzahlen berechnen ---
    total_xml = sum(len(info['rows']) for info in all_data.values())
    total_files = len(set(row.get('_filename', '') for info in all_data.values() for row in info['rows']))
    total_proto = sum(p['stats']['total'] for p in all_protocols) if all_protocols else 0
    total_missing = sum(len(v) for v in missing_by_file.values())
    total_matched = total_xml - total_missing
    total_proto_only = len(proto_only_keys)

    total_not_migrated = 0
    total_not_success = 0
    for p in all_protocols:
        total_not_migrated += p['stats']['not_migrated']
        total_not_success += p['stats']['not_success']

    # ═══════════════════════════════════════════════════════
    # BLOCK 1: ERGEBNIS-ZUSAMMENFASSUNG (ganz oben!)
    # ═══════════════════════════════════════════════════════
    r = 6
    if all_protocols:
        ws.cell(row=r, column=2, value="ERGEBNIS-ZUSAMMENFASSUNG").font = s.subtitle
        ws.merge_cells(f'B{r}:H{r}')
        r += 2

        # Grosse Status-Ampel
        all_ok = (total_missing == 0 and total_not_migrated == 0
                  and total_not_success == 0 and total_proto_only == 0)
        if all_ok:
            status_text = "ALLES OK - Alle Datensaetze erfolgreich migriert"
            status_font = s.success
            status_fill = s.success_bg
        else:
            status_text = "PRUEFUNG ERFORDERLICH - Abweichungen gefunden"
            status_font = s.error
            status_fill = s.error_bg

        cell = ws.cell(row=r, column=2, value=status_text)
        cell.font = Font(name='Arial', bold=True, color=EXCEL_GREEN if all_ok else EXCEL_RED, size=14)
        ws.merge_cells(f'B{r}:H{r}')
        for col in range(2, 9):
            ws.cell(row=r, column=col).fill = status_fill
        r += 2

        # Vergleich-Tabelle
        for col, h in enumerate(['', 'Pruefpunkt', 'Ergebnis', 'Status', '', '', ''], 2):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = s.header; cell.fill = s.dark_fill; cell.alignment = s.center; cell.border = s.thin
        r += 1

        checks = [
            ('Datensaetze in XML', str(total_xml), None, None),
            ('Datensaetze im Protokoll', str(total_proto), None, None),
            ('Zugeordnet (Match)', str(total_matched), total_matched == total_xml, 'success_bg' if total_matched == total_xml else 'error_bg'),
            ('Fehlend im Protokoll', str(total_missing), total_missing == 0, 'success_bg' if total_missing == 0 else 'error_bg'),
            ('Nur im Protokoll (nicht in XML)', str(total_proto_only), total_proto_only == 0, 'success_bg' if total_proto_only == 0 else 'warn_bg'),
            ('Nicht-Migriert (Action)', str(total_not_migrated), total_not_migrated == 0, 'success_bg' if total_not_migrated == 0 else 'error_bg'),
            ('Nicht-Success (Status)', str(total_not_success), total_not_success == 0, 'success_bg' if total_not_success == 0 else 'error_bg'),
        ]

        for desc, val, is_ok, fill_key in checks:
            status_icon = ''
            if is_ok is True: status_icon = 'OK'
            elif is_ok is False: status_icon = 'PRUEFEN'
            for col, cv in enumerate(['', desc, val, status_icon, '', '', ''], 2):
                cell = ws.cell(row=r, column=col, value=cv)
                cell.font = s.data; cell.alignment = s.center if col in (4,5) else s.left; cell.border = s.thin
                if col == 5 and fill_key:
                    cell.fill = getattr(s, fill_key)
                    cell.font = s.success if is_ok else (s.error if fill_key == 'error_bg' else s.warning)
            r += 1
        r += 1

        # ═══════════════════════════════════════════════════
        # BLOCK 2: INTELLIGENTE AUSWERTUNG + FRAGEN
        # ═══════════════════════════════════════════════════
        ws.cell(row=r, column=2, value="AUSWERTUNG & HANDLUNGSBEDARF").font = s.subtitle
        ws.merge_cells(f'B{r}:H{r}')
        r += 1

        findings = []
        questions = []

        if total_missing > 0:
            pct = round(total_missing / total_xml * 100, 1) if total_xml else 0
            findings.append(
                f"{total_missing} von {total_xml} Datensaetzen ({pct}%) aus der XML "
                f"sind NICHT im Migrationsprotokoll enthalten.")
            questions.append(
                f"Frage an Mandanten: Wurden diese {total_missing} Anlagen bewusst "
                f"nicht migriert oder fehlt ein weiteres Migrationsprotokoll?")

        if total_proto_only > 0:
            findings.append(
                f"{total_proto_only} Datensaetze im Protokoll haben KEINEN "
                f"passenden Eintrag in den XML-Dateien.")
            questions.append(
                f"Frage an Mandanten: Woher stammen diese {total_proto_only} "
                f"zusaetzlichen Datensaetze? Wurden Anlagen ausserhalb des "
                f"Migrationsobjekts migriert?")

        if total_proto > total_xml:
            diff = total_proto - total_xml
            findings.append(
                f"Das Protokoll enthaelt {diff} Datensaetze MEHR als die XML-Dateien "
                f"({total_proto} vs. {total_xml}).")
            questions.append(
                f"Frage an Mandanten: Sind zusaetzliche Anlagen migriert worden, "
                f"die nicht in der Sollmigrationsliste (XML) enthalten waren?")

        if total_not_migrated > 0:
            findings.append(
                f"{total_not_migrated} Datensaetze im Protokoll haben NICHT den "
                f"Action-Status 'Migrated/Migriert'.")
            questions.append(
                f"Frage an Mandanten: Warum wurden diese Datensaetze nicht migriert? "
                f"Gibt es technische Fehler oder bewusste Ausnahmen?")

        if total_not_success > 0:
            findings.append(
                f"{total_not_success} Datensaetze im Protokoll haben NICHT den "
                f"Status 'Success/Erfolg'.")
            questions.append(
                f"Frage an Mandanten: Welche Massnahmen wurden fuer die "
                f"fehlgeschlagenen Migrationen ergriffen?")

        if not findings:
            findings.append("Keine Abweichungen festgestellt. Alle Datensaetze wurden erfolgreich zugeordnet und migriert.")

        # Findings als Tabelle
        for col, h in enumerate(['', 'Nr.', 'Feststellung', '', '', '', ''], 2):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = s.header; cell.fill = s.orange_fill; cell.alignment = s.center; cell.border = s.thin
        r += 1
        for fi, finding in enumerate(findings, 1):
            ws.cell(row=r, column=3, value=fi).font = s.data_bold
            ws.cell(row=r, column=3).alignment = s.center; ws.cell(row=r, column=3).border = s.thin
            cell = ws.cell(row=r, column=4, value=finding)
            cell.font = s.info_text; cell.alignment = s.left_wrap; cell.border = s.thin
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
            ws.row_dimensions[r].height = 35
            r += 1
        r += 1

        # Fragen an Mandanten
        if questions:
            for col, h in enumerate(['', 'Nr.', 'Rueckfrage an den Mandanten', '', '', '', ''], 2):
                cell = ws.cell(row=r, column=col, value=h)
                cell.font = s.header; cell.fill = PatternFill(start_color=EXCEL_ORANGE_LIGHT, end_color=EXCEL_ORANGE_LIGHT, patternType='solid')
                cell.alignment = s.center; cell.border = s.thin
            r += 1
            for qi, q in enumerate(questions, 1):
                ws.cell(row=r, column=3, value=qi).font = s.data_bold
                ws.cell(row=r, column=3).alignment = s.center; ws.cell(row=r, column=3).border = s.thin
                cell = ws.cell(row=r, column=4, value=q)
                cell.font = s.question; cell.alignment = s.left_wrap; cell.border = s.thin
                ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
                for col in range(3, 9):
                    ws.cell(row=r, column=col).fill = s.question_bg
                ws.row_dimensions[r].height = 40
                r += 1
            r += 1

    else:
        # Kein Protokoll → Nur KPIs
        r = 6
        ws.cell(row=r, column=2, value="KENNZAHLEN").font = s.subtitle
        ws.merge_cells(f'B{r}:H{r}')
        r += 2

    # ═══════════════════════════════════════════════════════
    # BLOCK 3: KPI-Kaesten
    # ═══════════════════════════════════════════════════════
    ws.cell(row=r, column=2, value="KENNZAHLEN").font = s.subtitle
    ws.merge_cells(f'B{r}:H{r}')
    r += 2
    kpis = [('B', 'C', str(total_files), 'XML Dateien'),
            ('D', 'E', str(total_xml), 'Datensaetze XML'),
            ('F', 'G', str(total_proto), 'Saetze Protokoll'),
            ('H', 'I', str(len(all_protocols)), 'Protokolle')]
    for c1l, c2l, val, label in kpis:
        c1, c2 = ord(c1l)-64, ord(c2l)-64
        cell = ws.cell(row=r, column=c1, value=val)
        cell.font = s.kpi_value; cell.alignment = s.center; cell.fill = s.kpi_fill; cell.border = s.kpi_border
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
        cell = ws.cell(row=r+1, column=c1, value=label)
        cell.font = s.kpi_label; cell.alignment = s.center; cell.fill = s.kpi_fill
        ws.merge_cells(start_row=r+1, start_column=c1, end_row=r+1, end_column=c2)
    r += 3

    # ═══════════════════════════════════════════════════════
    # BLOCK 4: Protokoll-Analyse
    # ═══════════════════════════════════════════════════════
    if all_protocols:
        ws.cell(row=r, column=2, value="PROTOKOLL-ANALYSE").font = s.subtitle
        ws.merge_cells(f'B{r}:H{r}')
        r += 2
        for col, h in enumerate(['', 'Protokoll-Datei', 'Saetze', 'Migrated', 'Success', 'Status', ''], 2):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = s.header; cell.fill = s.dark_fill; cell.alignment = s.center; cell.border = s.thin
        r += 1
        for proto in all_protocols:
            st = proto['stats']
            is_ok = st['all_migrated'] and st['all_success']
            for col, val in enumerate(['', proto['filename'], st['total'],
                                       f"{st['migrated']}/{st['total']}", f"{st['success']}/{st['total']}",
                                       'OK' if is_ok else 'PRUEFEN', ''], 2):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = s.data; cell.alignment = s.center; cell.border = s.thin
                if col == 7:
                    cell.fill = s.success_bg if is_ok else s.error_bg
                    cell.font = s.success if is_ok else s.error
            r += 1
        r += 1

    # ═══════════════════════════════════════════════════════
    # BLOCK 5: Fehlende Datensaetze je Datei
    # ═══════════════════════════════════════════════════════
    if missing_by_file:
        ws.cell(row=r, column=2, value="FEHLENDE DATENSAETZE JE DATEI").font = s.subtitle
        ws.merge_cells(f'B{r}:H{r}')
        r += 1
        for col, h in enumerate(['', 'XML-Dateiname', 'Anzahl fehlend', 'Beispiele (Match-Keys)', '', '', ''], 2):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = s.header; cell.fill = s.red_fill; cell.alignment = s.center; cell.border = s.thin
        r += 1
        for fname, keys in missing_by_file.items():
            ex = ", ".join(keys[:5])
            if len(keys) > 5:
                ex += f" ... (+{len(keys)-5} weitere)"
            for col, val in enumerate(['', fname, len(keys), ex, '', '', ''], 2):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = s.data; cell.alignment = s.left; cell.border = s.thin; cell.fill = s.error_bg
            ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=8)
            r += 1
        r += 1

    # ═══════════════════════════════════════════════════════
    # BLOCK 6: Datei-Uebersicht
    # ═══════════════════════════════════════════════════════
    ws.cell(row=r, column=2, value="DATEI-UEBERSICHT").font = s.subtitle
    ws.merge_cells(f'B{r}:H{r}')
    r += 2
    sn = 0
    for key_sig, info in all_data.items():
        sn += 1
        ws.cell(row=r, column=2, value=f"Sheet: Key_{sn}").font = s.data_bold
        ws.cell(row=r, column=4, value=f"Key: {' | '.join(info['fields'])}").font = Font(name='Arial', italic=True, color='666666', size=10)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=8)
        r += 1
        for col, h in enumerate(['', 'Nr.', 'Dateiname', 'Datensaetze', 'Sheet', ''], 2):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = s.header; cell.fill = s.orange_fill; cell.alignment = s.center; cell.border = s.thin
        r += 1
        fc = {}
        for rd in info['rows']:
            fn = rd.get('_filename', '')
            fc[fn] = fc.get(fn, 0) + 1
        nr = 0; tot = 0
        for fn, cnt in fc.items():
            nr += 1; tot += cnt
            for col, val in enumerate(['', nr, fn, cnt, f'Key_{sn}', ''], 2):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = s.data; cell.alignment = s.center; cell.border = s.thin
                if r % 2 == 0: cell.fill = s.even_fill
            r += 1
        for col, val in enumerate(['', '', 'GESAMT', tot, '', ''], 2):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = s.total; cell.border = s.thin; cell.fill = s.total_fill
        r += 2

    # ═══════════════════════════════════════════════════════
    # BLOCK 7: VERGLEICH PRO KEY-TYP
    # ═══════════════════════════════════════════════════════
    if per_key_stats and all_protocols:
        ws.cell(row=r, column=2, value="VERGLEICH PRO KEY-TYP").font = s.subtitle
        ws.merge_cells(f'B{r}:H{r}')
        r += 1
        ws.cell(row=r, column=2, value="Jeder Key-Typ wird separat gegen das Protokoll verglichen (siehe Vergl_1, Vergl_2, ...)").font = s.formula_hint
        ws.merge_cells(f'B{r}:H{r}')
        r += 1
        for col, h in enumerate(['', 'Key-Typ', 'XML Zeilen', 'Unique Keys', 'Matched', 'Fehlend', 'Duplikate'], 2):
            cell = ws.cell(row=r, column=col, value=h)
            cell.font = s.header; cell.fill = s.dark_fill; cell.alignment = s.center; cell.border = s.thin
        r += 1
        for key_sig in all_data:
            ks = per_key_stats.get(key_sig, {})
            kn = ks.get('key_num', '?')
            key_label = ' | '.join(all_data[key_sig]['fields'])
            is_ok = ks.get('missing', 0) == 0
            for col, val in enumerate(['', f'Key_{kn} ({key_label})', ks.get('total_rows', 0),
                                        ks.get('unique_keys', 0), ks.get('matched', 0),
                                        ks.get('missing', 0), ks.get('duplicate_count', 0)], 2):
                cell = ws.cell(row=r, column=col, value=val)
                cell.font = s.data; cell.alignment = s.center; cell.border = s.thin
                if col == 7:
                    cell.fill = s.success_bg if is_ok else s.error_bg
                    cell.font = s.success if is_ok else s.error
            r += 1
        r += 1

    ws.column_dimensions['A'].width = 3; ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22; ws.column_dimensions['D'].width = 42
    ws.column_dimensions['E'].width = 20; ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 16; ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 3
    ws.sheet_properties.tabColor = EXCEL_ORANGE


def _create_anleitung(wb, s, all_data=None):
    """Erklaert was geprueft wird, welche Formeln verwendet werden, und die Sheet-Struktur."""
    ws = wb.create_sheet(title="Anleitung")
    ws.sheet_properties.tabColor = EXCEL_DARK

    for row in range(1, 80):
        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = s.bg_fill

    for col in range(1, 8):
        ws.cell(row=1, column=col).fill = s.orange_fill
    ws.cell(row=2, column=2, value="Anleitung & Formeldokumentation").font = s.title
    ws.merge_cells('B2:F2')
    for col in range(1, 8):
        ws.cell(row=3, column=col).border = Border(bottom=Side(style='medium', color=EXCEL_ORANGE))

    r = 5
    # --- Was wird geprueft? ---
    ws.cell(row=r, column=2, value="WAS WIRD GEPRUEFT?").font = s.subtitle
    ws.merge_cells(f'B{r}:F{r}')
    r += 2

    pruefungen = [
        ("Vollstaendigkeit",
         "Jeder Datensatz aus den SAP-Migrations-XML-Dateien wird im Migrationsprotokoll (TXT) gesucht. "
         "Fehlende Datensaetze werden als 'FEHLT' markiert."),
        ("Migrationsstatus",
         "Fuer jeden gefundenen Datensatz wird geprueft, ob der Action-Status 'Migrated' (bzw. 'Migriert') ist. "
         "Abweichungen werden rot hervorgehoben."),
        ("Erfolgsstatus",
         "Fuer jeden gefundenen Datensatz wird geprueft, ob der Status 'Success' (bzw. 'Erfolg') ist. "
         "Abweichungen werden rot hervorgehoben."),
        ("Ueberhaenge im Protokoll",
         "Datensaetze die NUR im Protokoll stehen aber nicht in der XML enthalten sind, werden gezaehlt. "
         "Dies kann auf zusaetzliche Migrationen hindeuten."),
        ("Zusammengefuehrter Schluessel",
         "Die Zuordnung erfolgt ueber einen zusammengesetzten Schluessel (Match_Key) aus den Key-Feldern "
         "des Migrationsobjekts (z.B. Buchungskreis|Anlagennummer|Unternummer). "
         "Dadurch ist die Zuordnung unabhaengig von der Spaltenreihenfolge."),
    ]

    for col, h in enumerate(['', 'Nr.', 'Pruefung', 'Beschreibung', '', ''], 2):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = s.header; cell.fill = s.dark_fill; cell.alignment = s.center; cell.border = s.thin
    r += 1
    for pi, (title, desc) in enumerate(pruefungen, 1):
        ws.cell(row=r, column=3, value=pi).font = s.data_bold
        ws.cell(row=r, column=3).alignment = s.center; ws.cell(row=r, column=3).border = s.thin
        ws.cell(row=r, column=4, value=title).font = s.data_bold
        ws.cell(row=r, column=4).border = s.thin
        cell = ws.cell(row=r, column=5, value=desc)
        cell.font = s.info_text; cell.alignment = s.left_wrap; cell.border = s.thin
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 40
        r += 1
    r += 2

    # --- Sheet-Uebersicht ---
    ws.cell(row=r, column=2, value="SHEET-UEBERSICHT").font = s.subtitle
    ws.merge_cells(f'B{r}:F{r}')
    r += 2

    sheets = [
        ("Uebersicht", "Zusammenfassung aller Ergebnisse, Kennzahlen, Auswertung, Rueckfragen an den Mandanten."),
        ("Anleitung", "Diese Seite. Dokumentation aller Pruefungen und Formeln."),
        ("Protokoll / Proto_X", "Rohdaten des Migrationsprotokolls. Spalte A = Match_Key, B = Action (normalisiert), C = Status (normalisiert)."),
        ("Key_X", "Rohdaten der XML-Dateien mit allen Key-Feldern und dem Match_Key."),
        ("Vergl_1, Vergl_2, ...", "Pro Key-Typ ein separater Kreuzvergleich XML vs. Protokoll. Match ueber erstes Key-Feld. INDEX/MATCH-Formeln, editierbar in Excel. Bei Key-Typen mit Duplikaten wird die Anzahl XML-Zeilen pro Key angezeigt."),
        ("Fehlerhafte Dateien", "XML-Dateien die nicht gelesen werden konnten (nur wenn Fehler auftraten)."),
    ]

    for col, h in enumerate(['', 'Nr.', 'Sheet', 'Beschreibung', '', ''], 2):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = s.header; cell.fill = s.dark_fill; cell.alignment = s.center; cell.border = s.thin
    r += 1
    for si, (name, desc) in enumerate(sheets, 1):
        ws.cell(row=r, column=3, value=si).font = s.data_bold
        ws.cell(row=r, column=3).alignment = s.center; ws.cell(row=r, column=3).border = s.thin
        ws.cell(row=r, column=4, value=name).font = s.data_bold
        ws.cell(row=r, column=4).border = s.thin
        cell = ws.cell(row=r, column=5, value=desc)
        cell.font = s.info_text; cell.alignment = s.left_wrap; cell.border = s.thin
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 35
        r += 1
    r += 2

    # --- FORMEL-DOKUMENTATION ---
    ws.cell(row=r, column=2, value="FORMEL-DOKUMENTATION (Vergleich-Sheet)").font = s.subtitle
    ws.merge_cells(f'B{r}:F{r}')
    r += 2

    formeln = [
        ("Spalte C: Action",
         '\'=IFERROR(INDEX(Protokoll!$B$2:$B$N, MATCH(A5, Protokoll!$A$2:$A$N, 0)), "NICHT GEFUNDEN")',
         "Sucht den Match_Key (Spalte A) im Protokoll-Sheet und gibt den normalisierten Action-Wert zurueck. "
         "'NICHT GEFUNDEN' wenn kein Match."),
        ("Spalte D: Status",
         '\'=IFERROR(INDEX(Protokoll!$C$2:$C$N, MATCH(A5, Protokoll!$A$2:$A$N, 0)), "NICHT GEFUNDEN")',
         "Sucht den Match_Key im Protokoll-Sheet und gibt den normalisierten Status-Wert zurueck."),
        ("Spalte E: Alles OK?",
         '\'=IF(OR(C5="NICHT GEFUNDEN",D5="NICHT GEFUNDEN"),"FEHLT", IF(AND(C5="Migrated",D5="Success"),"OK","FEHLER"))',
         "Kombinierte Pruefung: "
         "FEHLT = Datensatz nicht im Protokoll, "
         "OK = Action=Migrated UND Status=Success, "
         "FEHLER = Datensatz gefunden aber Action oder Status stimmt nicht."),
        ("Zusammenfassung: OK",
         '\'=COUNTIF(E5:E{n},"OK")',
         "Zaehlt alle Datensaetze mit Status OK."),
        ("Zusammenfassung: Fehler",
         '\'=COUNTIF(E5:E{n},"FEHLER")',
         "Zaehlt alle Datensaetze mit abweichendem Migrationsstatus."),
        ("Zusammenfassung: Fehlt",
         '\'=COUNTIF(E5:E{n},"FEHLT")',
         "Zaehlt alle Datensaetze die nicht im Protokoll gefunden wurden."),
    ]

    for col, h in enumerate(['', 'Nr.', 'Formel-Name', 'Formel (Beispiel)', 'Erklaerung'], 2):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = s.header; cell.fill = s.orange_fill; cell.alignment = s.center; cell.border = s.thin
    r += 1
    for fi, (name, formula, desc) in enumerate(formeln, 1):
        ws.cell(row=r, column=3, value=fi).font = s.data_bold
        ws.cell(row=r, column=3).alignment = s.center; ws.cell(row=r, column=3).border = s.thin
        ws.cell(row=r, column=4, value=name).font = s.data_bold
        ws.cell(row=r, column=4).border = s.thin
        cell = ws.cell(row=r, column=5, value=formula)
        cell.font = Font(name='Consolas', size=9, color='333333')
        cell.alignment = s.left_wrap; cell.border = s.thin
        cell = ws.cell(row=r, column=6, value=desc)
        cell.font = s.info_text; cell.alignment = s.left_wrap; cell.border = s.thin
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
        ws.row_dimensions[r].height = 50
        r += 1
    r += 2

    # --- Normalisierung ---
    ws.cell(row=r, column=2, value="NORMALISIERUNG (Deutsch/Englisch)").font = s.subtitle
    ws.merge_cells(f'B{r}:F{r}')
    r += 2

    for col, h in enumerate(['', '', 'Deutsch (Original)', 'Englisch (Normalisiert)', 'Verwendet in'], 2):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = s.header; cell.fill = s.dark_fill; cell.alignment = s.center; cell.border = s.thin
    r += 1
    for de, en, used in [('Migriert', 'Migrated', 'Spalte C (Action)'),
                          ('Erfolg', 'Success', 'Spalte D (Status)'),
                          ('Feldliste', 'Field List', 'XML Sheet-Name'),
                          ('Schluessel', 'Key', 'XML Gruppen-Name'),
                          ('(obligatorisch)', '(mandatory)', 'XML Sheet-Marker')]:
        for col, val in enumerate(['', '', de, en, used], 2):
            cell = ws.cell(row=r, column=col, value=val)
            cell.font = s.data; cell.alignment = s.center; cell.border = s.thin
            if r % 2 == 0: cell.fill = s.even_fill
        r += 1

    ws.column_dimensions['A'].width = 3; ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 8; ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 55; ws.column_dimensions['F'].width = 45
    ws.column_dimensions['G'].width = 25


def _create_protocol_sheet(wb, proto, s, sheet_name):
    data_rows = proto['data']
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = EXCEL_DARK
    proto_headers = proto.get('headers', [])
    display_headers = ['Match_Key', 'Action_Norm', 'Status_Norm'] + proto_headers
    for col, h in enumerate(display_headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = s.header; cell.fill = s.dark_fill; cell.alignment = s.center; cell.border = s.thin
    for ri, rd in enumerate(data_rows, 2):
        ws.cell(row=ri, column=1, value=rd.get('_match_key', '')).font = s.data
        ws.cell(row=ri, column=1).alignment = s.left; ws.cell(row=ri, column=1).border = s.thin
        ws.cell(row=ri, column=2, value=rd.get('Action_Normalized', '')).font = s.data
        ws.cell(row=ri, column=2).alignment = s.center; ws.cell(row=ri, column=2).border = s.thin
        ws.cell(row=ri, column=3, value=rd.get('Status_Normalized', '')).font = s.data
        ws.cell(row=ri, column=3).alignment = s.center; ws.cell(row=ri, column=3).border = s.thin
        for ci, h in enumerate(proto_headers, 4):
            cell = ws.cell(row=ri, column=ci, value=rd.get(h, ''))
            cell.font = s.data; cell.alignment = s.center if h in ('Action', 'Status') else s.left
            cell.border = s.thin
        if ri % 2 == 0:
            for ci in range(1, len(display_headers)+1):
                ws.cell(row=ri, column=ci).fill = s.even_fill
    lr = len(data_rows) + 1
    if lr >= 2:
        ws.conditional_formatting.add(f'B2:B{lr}',
            CellIsRule(operator='notEqual', formula=['"Migrated"'],
                fill=PatternFill(start_color=EXCEL_ERROR_BG, end_color=EXCEL_ERROR_BG, patternType='solid'),
                font=Font(color=EXCEL_RED, bold=True)))
        ws.conditional_formatting.add(f'C2:C{lr}',
            CellIsRule(operator='notEqual', formula=['"Success"'],
                fill=PatternFill(start_color=EXCEL_ERROR_BG, end_color=EXCEL_ERROR_BG, patternType='solid'),
                font=Font(color=EXCEL_RED, bold=True)))
        ws.auto_filter.ref = f'A1:{get_column_letter(len(display_headers))}{lr}'
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    for ci in range(4, len(display_headers)+1):
        ws.column_dimensions[get_column_letter(ci)].width = 22


def _create_data_sheets(wb, all_data, s):
    sn = 0
    for key_sig, info in all_data.items():
        sn += 1
        fields = info['fields']; rows = info['rows']
        ws = wb.create_sheet(title=f"Key_{sn}")
        ws.sheet_properties.tabColor = EXCEL_ORANGE_LIGHT
        headers = fields + ['Match_Key', 'Dateiname']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = s.header; cell.fill = s.orange_fill; cell.alignment = s.center; cell.border = s.thin
        for ri, rd in enumerate(rows, 2):
            for ci, f in enumerate(fields, 1):
                cell = ws.cell(row=ri, column=ci, value=rd.get(f, ''))
                cell.font = s.data; cell.alignment = s.left; cell.border = s.thin
                if ri % 2 == 0: cell.fill = s.even_fill
            cell = ws.cell(row=ri, column=len(fields)+1, value=rd.get('_match_key', ''))
            cell.font = s.data; cell.alignment = s.left; cell.border = s.thin
            if ri % 2 == 0: cell.fill = s.even_fill
            cell = ws.cell(row=ri, column=len(fields)+2, value=rd.get('_filename', ''))
            cell.font = s.data; cell.alignment = s.left; cell.border = s.thin
            if ri % 2 == 0: cell.fill = s.even_fill
        for ci in range(1, len(headers)+1):
            cl = get_column_letter(ci)
            mx = len(str(headers[ci-1]))
            for ri2 in range(2, min(len(rows)+2, 100)):
                v = ws.cell(row=ri2, column=ci).value
                if v: mx = max(mx, len(str(v)))
            ws.column_dimensions[cl].width = min(mx+4, 45)
        ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}1'


def _create_comparison_sheet(wb, info, key_num, s, proto_sheet, proto, sheet_name, key_stats):
    """Erstellt ein Vergleich-Sheet fuer EINEN Key-Typ gegen das Protokoll.
    Match erfolgt ueber das ERSTE Key-Feld (Product Number / Material Number)."""
    ws = wb.create_sheet(title=sheet_name)
    ws.sheet_properties.tabColor = EXCEL_GREEN
    fields = info['fields']
    first_field = fields[0]
    has_dupes = key_stats.get('has_duplicates', False)
    dupe_count = key_stats.get('duplicate_count', 0)

    # Unique Keys sammeln (nur erstes Feld = Match gegen Protokoll)
    all_keys = []
    key_fnames = {}
    key_display = {}
    key_row_count = {}
    for row in info['rows']:
        pk = row.get(first_field, '').strip()
        if pk:
            if pk not in key_fnames:
                all_keys.append(pk)
                key_fnames[pk] = row.get('_filename', '')
                key_display[pk] = ' | '.join(row.get(f, '') for f in fields)
                key_row_count[pk] = 1
            else:
                key_row_count[pk] = key_row_count.get(pk, 1) + 1

    # Header
    key_label = ' | '.join(fields)
    ws.cell(row=1, column=1, value=f"Vergleich Key_{key_num}: XML vs. {proto['filename']}").font = s.subtitle
    ws.merge_cells('A1:G1')
    hint = f"Key-Felder: {key_label} | Match ueber: {first_field} | Spalten C-D = INDEX/MATCH Formeln"
    if has_dupes:
        hint += f" | HINWEIS: {dupe_count} Duplikate (Spalte G zeigt Anzahl XML-Zeilen pro Key)"
    ws.cell(row=2, column=1, value=hint).font = s.formula_hint
    ws.merge_cells('A2:G2')

    headers = ['Match_Key', 'Key (Klartext)', 'Action', 'Status', 'Alles OK?', 'Dateiname']
    if has_dupes:
        headers.append('Anz. XML-Zeilen')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = s.header; cell.fill = s.orange_fill; cell.alignment = s.center; cell.border = s.thin

    plast = len(proto['data']) + 1

    for ri, mk in enumerate(all_keys, 5):
        ws.cell(row=ri, column=1, value=mk).font = s.data
        ws.cell(row=ri, column=1).alignment = s.left; ws.cell(row=ri, column=1).border = s.thin

        ws.cell(row=ri, column=2, value=key_display.get(mk, '')).font = s.data
        ws.cell(row=ri, column=2).alignment = s.left; ws.cell(row=ri, column=2).border = s.thin

        cell = ws.cell(row=ri, column=3)
        cell.value = f'=IFERROR(INDEX(\'{proto_sheet}\'!$B$2:$B${plast},MATCH(A{ri},\'{proto_sheet}\'!$A$2:$A${plast},0)),"NICHT GEFUNDEN")'
        cell.font = s.data; cell.alignment = s.center; cell.border = s.thin

        cell = ws.cell(row=ri, column=4)
        cell.value = f'=IFERROR(INDEX(\'{proto_sheet}\'!$C$2:$C${plast},MATCH(A{ri},\'{proto_sheet}\'!$A$2:$A${plast},0)),"NICHT GEFUNDEN")'
        cell.font = s.data; cell.alignment = s.center; cell.border = s.thin

        cell = ws.cell(row=ri, column=5)
        cell.value = f'=IF(OR(C{ri}="NICHT GEFUNDEN",D{ri}="NICHT GEFUNDEN"),"FEHLT",IF(AND(C{ri}="Migrated",D{ri}="Success"),"OK","FEHLER"))'
        cell.font = s.data_bold; cell.alignment = s.center; cell.border = s.thin

        ws.cell(row=ri, column=6, value=key_fnames.get(mk, '')).font = s.data
        ws.cell(row=ri, column=6).alignment = s.left; ws.cell(row=ri, column=6).border = s.thin

        if has_dupes:
            cnt = key_row_count.get(mk, 1)
            cell = ws.cell(row=ri, column=7, value=cnt)
            cell.font = s.data_bold if cnt > 1 else s.data
            cell.alignment = s.center; cell.border = s.thin

    ld = len(all_keys) + 4

    if ld >= 5:
        ws.conditional_formatting.add(f'E5:E{ld}',
            CellIsRule(operator='equal', formula=['"OK"'],
                fill=PatternFill(start_color=EXCEL_SUCCESS_BG, end_color=EXCEL_SUCCESS_BG, patternType='solid'),
                font=Font(color=EXCEL_GREEN, bold=True)))
        ws.conditional_formatting.add(f'E5:E{ld}',
            CellIsRule(operator='equal', formula=['"FEHLER"'],
                fill=PatternFill(start_color=EXCEL_ERROR_BG, end_color=EXCEL_ERROR_BG, patternType='solid'),
                font=Font(color=EXCEL_RED, bold=True)))
        ws.conditional_formatting.add(f'E5:E{ld}',
            CellIsRule(operator='equal', formula=['"FEHLT"'],
                fill=PatternFill(start_color=EXCEL_WARN_BG, end_color=EXCEL_WARN_BG, patternType='solid'),
                font=Font(color=EXCEL_YELLOW, bold=True)))
        ws.conditional_formatting.add(f'C5:C{ld}',
            CellIsRule(operator='notEqual', formula=['"Migrated"'],
                fill=PatternFill(start_color=EXCEL_ERROR_BG, end_color=EXCEL_ERROR_BG, patternType='solid'),
                font=Font(color=EXCEL_RED, bold=True)))

    sr = ld + 2
    ws.cell(row=sr, column=1, value="ZUSAMMENFASSUNG").font = s.subtitle
    ws.merge_cells(f'A{sr}:F{sr}')
    sr += 1
    ws.cell(row=sr, column=1, value=f"Key-Typ:").font = s.data_bold
    ws.cell(row=sr, column=2, value=f"Key_{key_num} ({key_label})").font = s.data
    sr += 1
    ws.cell(row=sr, column=1, value="Unique Keys:").font = s.data_bold
    ws.cell(row=sr, column=2, value=len(all_keys)).font = s.data_bold
    if has_dupes:
        ws.cell(row=sr, column=3, value=f"({len(info['rows'])} Zeilen total, {dupe_count} Duplikate)").font = s.formula_hint
    sr += 1
    ws.cell(row=sr, column=1, value="OK:").font = s.data_bold
    ws.cell(row=sr, column=2).value = f'=COUNTIF(E5:E{ld},"OK")'
    ws.cell(row=sr, column=2).font = s.success
    sr += 1
    ws.cell(row=sr, column=1, value="Fehler:").font = s.data_bold
    ws.cell(row=sr, column=2).value = f'=COUNTIF(E5:E{ld},"FEHLER")'
    ws.cell(row=sr, column=2).font = s.error
    sr += 1
    ws.cell(row=sr, column=1, value="Fehlt:").font = s.data_bold
    ws.cell(row=sr, column=2).value = f'=COUNTIF(E5:E{ld},"FEHLT")'
    ws.cell(row=sr, column=2).font = s.warning

    ws.column_dimensions['A'].width = 30; ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18; ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14; ws.column_dimensions['F'].width = 35
    if has_dupes:
        ws.column_dimensions['G'].width = 16


def _create_error_sheet(wb, skipped_files, s):
    ws = wb.create_sheet(title="Fehlerhafte Dateien")
    ws.sheet_properties.tabColor = EXCEL_RED
    for col, h in enumerate(['Dateiname', 'Fehlermeldung'], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = s.header
        cell.fill = PatternFill(start_color=EXCEL_RED, end_color=EXCEL_RED, patternType='solid')
        cell.alignment = s.center; cell.border = s.thin
    for ri, (fn, err) in enumerate(skipped_files, 2):
        ws.cell(row=ri, column=1, value=fn).font = s.data
        ws.cell(row=ri, column=1).border = s.thin
        ws.cell(row=ri, column=2, value=err).font = s.data
        ws.cell(row=ri, column=2).border = s.thin
        ws.cell(row=ri, column=2).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 80